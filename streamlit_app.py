import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
from openpyxl.styles import Border, Side, Alignment, Font

st.set_page_config(page_title="PDF to Master Summary", layout="wide")

st.title("🚀 Tool Chuyển Đổi PDF sang Bảng Tổng Hợp Master")
st.markdown("""
- **Trang 1 (mỗi file):** Bỏ qua (Purchase Note).
- **Các trang còn lại:** Tự động lấy SO, Ngày tháng và **Tên Store (Cột C)**.
- **Có thể tải lên NHIỀU file PDF cùng lúc** — dữ liệu tất cả các file sẽ được **gộp chung vào 1 file Excel duy nhất**.
""")

uploaded_files = st.file_uploader(
    "Tải một hoặc nhiều file PDF lên",
    type="pdf",
    accept_multiple_files=True
)

if uploaded_files:
    with st.spinner(f'Đang phân tích {len(uploaded_files)} file PDF...'):
        all_rows = []
        total_pages_processed = 0

        for uploaded_file in uploaded_files:
            with pdfplumber.open(uploaded_file) as pdf:
                # Duyệt từ trang 2 (index 1) đến hết trang cuối của file này
                for i in range(1, len(pdf.pages)):
                    page = pdf.pages[i]
                    total_pages_processed += 1

                    # 1. Lấy thông tin chung (SO, Order Date, Delivery Date) bằng văn bản
                    text = page.extract_text() or ""
                    so_match = re.search(r"SO number:\s*(\d+)", text)
                    order_date = re.search(r"Order date:\s*([\d/ :]+)", text)
                    delivery_date = re.search(r"Delivery date:\s*([\d/]+)", text)

                    so = so_match.group(1) if so_match else ""
                    o_date = order_date.group(1).split()[0] if order_date else ""
                    d_date = delivery_date.group(1) if delivery_date else ""

                    # 2. LẤY TÊN STORE CHUẨN XÁC TẠI CỘT C
                    store_name = ""
                    # Trích xuất bảng đầu tiên (chứa địa chỉ)
                    header_tables = page.extract_tables()
                    if header_tables:
                        # Thông thường bảng Header là bảng đầu tiên của trang
                        header_data = header_tables[0]
                        # Dòng 3 (index 2), Cột C (index 2) là tên Store dưới chữ "For Store"
                        if len(header_data) >= 3 and len(header_data[2]) >= 3:
                            store_name = str(header_data[2][2]).replace('\n', ' ').strip()

                    # 3. Trích xuất bảng chi tiết sản phẩm
                    # Tìm tất cả các bảng và lọc bảng có chứa mã hàng
                    for table in header_tables:
                        for row in table:
                            clean_row = [str(cell).replace('\n', ' ') if cell else "" for cell in row]

                            # Điều kiện: Cột 1 là mã hàng (số) và độ dài >= 10
                            article_id = clean_row[0].strip()
                            if article_id.isdigit() and len(article_id) >= 10:
                                all_rows.append({
                                    "Source File": uploaded_file.name,
                                    "SO Number": so,
                                    "Order Date": o_date,
                                    "Delivery Date": d_date,
                                    "Store Name": store_name,  # Cột trong Excel kết quả
                                    "Article": article_id,
                                    "Description": clean_row[1],
                                    "OU Qty": clean_row[5] if len(clean_row) > 5 else ""
                                })

        if all_rows:
            # Tạo DataFrame
            df_final = pd.DataFrame(all_rows)
            
            # Chuyển đổi số lượng sang kiểu số để tính toán nếu cần
            df_final["OU Qty"] = pd.to_numeric(df_final["OU Qty"], errors='coerce')
            
            # Tạo bảng Pivot: Tổng số lượng theo Store x SKU (Article + Description)
            pivot_df = pd.pivot_table(
                df_final,
                index="Store Name",
                columns=["Article", "Description"],
                values="OU Qty",
                aggfunc="sum",
                fill_value=0,
            )
            # Gộp tên cột MultiIndex (Article, Description) thành 1 chuỗi dễ đọc
            pivot_df.columns = [f"{article} - {desc}" for article, desc in pivot_df.columns]
            pivot_df["TOTAL"] = pivot_df.sum(axis=1)
            pivot_df = pivot_df.reset_index()

            # Ghi ra file Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name="Master_Summary")
                pivot_df.to_excel(writer, index=False, sheet_name="Pivot_Store_by_SKU")

                thin = Side(style='thin')
                border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # Định dạng sheet Master_Summary (Đóng khung + Căn giữa Header)
                ws = writer.sheets["Master_Summary"]
                for row in ws.iter_rows(min_row=1, max_row=len(all_rows)+1, max_col=len(df_final.columns)):
                    for cell in row:
                        cell.border = border
                        if cell.row == 1:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")

                # Định dạng sheet Pivot_Store_by_SKU
                ws_pivot = writer.sheets["Pivot_Store_by_SKU"]
                for row in ws_pivot.iter_rows(min_row=1, max_row=len(pivot_df)+1, max_col=len(pivot_df.columns)):
                    for cell in row:
                        cell.border = border
                        if cell.row == 1:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")
                # Bôi đậm cột TOTAL cho dễ nhìn
                total_col_idx = list(pivot_df.columns).index("TOTAL") + 1
                for row in ws_pivot.iter_rows(min_row=2, max_row=len(pivot_df)+1,
                                               min_col=total_col_idx, max_col=total_col_idx):
                    for cell in row:
                        cell.font = Font(bold=True)

            st.success(
                f"✅ Đã xử lý thành công {len(uploaded_files)} file PDF "
                f"({total_pages_processed} trang, {len(all_rows)} dòng dữ liệu)!"
            )

            st.subheader("📋 Sheet 1: Master_Summary (chi tiết từng dòng)")
            st.dataframe(df_final)

            st.subheader("📊 Sheet 2: Pivot_Store_by_SKU (tổng số lượng theo Store x SKU)")
            st.dataframe(pivot_df)

            st.download_button(
                label="📥 Tải file Excel Tổng Hợp (2 sheet: chi tiết + pivot theo Store/SKU)",
                data=output.getvalue(),
                file_name="Bao_cao_Tong_Hop_Master.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("⚠️ Không tìm thấy dữ liệu phù hợp trong (các) file PDF.")